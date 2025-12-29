import requests
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===================== CONFIG =====================
JIRA_URL = "https://jira.demo.almworks.com"
TOKEN = "NzAyNDQ5MTg0MTE4Ou3Ke9W2Wq9URNolZRMRtZEafIQk"

HEADERS = {
    "Authorization": f"Bearer {TOKEN}",
    "Content-Type": "application/json"
}

SEARCH_API = f"{JIRA_URL}/rest/api/2/search"

# ===================== JQL DEFINITIONS =====================
JQLS = {
    "Accepted": (
        "filter=11400 AND labels = AppSecurity "
        "AND labels = appsec_validated "
        "AND labels = appsec_approved"
    ),
    "Rejected": (
        "filter=11400 AND labels = AppSecurity "
        "AND labels = appsec_validated "
        "AND labels = appsec_rejected"
    ),
    "Doable": (
        "filter=11400 AND labels = AppSecurity "
        "AND labels not in (appsec_validated) "
        "AND labels not in (appsec_approved, appsec_rejected) "
        "AND status in (Done, Resolved)"
    ),
    "NonDoable": (
        "filter=11400 AND labels = AppSecurity "
        "AND labels not in (appsec_validated) "
        "AND labels not in (appsec_approved, appsec_rejected) "
        "AND status not in (Done, Resolved)"
    )
}

# ===================== FETCH ISSUE COUNTS =====================
def fetch_issue_type_counts(jql):
    counts = defaultdict(int)
    start_at = 0
    max_results = 100

    while True:
        payload = {
            "jql": jql,
            "startAt": start_at,
            "maxResults": max_results,
            "fields": ["issuetype"]
        }

        response = requests.post(SEARCH_API, headers=HEADERS, json=payload)
        response.raise_for_status()
        data = response.json()

        for issue in data.get("issues", []):
            issue_type = issue["fields"]["issuetype"]["name"]
            counts[issue_type] += 1

        if start_at + max_results >= data.get("total", 0):
            break

        start_at += max_results

    return dict(counts)

# ===================== FETCH DATA =====================
accepted = fetch_issue_type_counts(JQLS["Accepted"])
rejected = fetch_issue_type_counts(JQLS["Rejected"])
doable = fetch_issue_type_counts(JQLS["Doable"])
nondoable = fetch_issue_type_counts(JQLS["NonDoable"])

all_issue_types = sorted(
    set(accepted) | set(rejected) | set(doable) | set(nondoable)
)

reviewed = {k: accepted.get(k, 0) + rejected.get(k, 0) for k in all_issue_types}
pending = {k: doable.get(k, 0) + nondoable.get(k, 0) for k in all_issue_types}
overall = {k: reviewed[k] + pending[k] for k in all_issue_types}

# ===================== EXCEL =====================
wb = Workbook()
ws = wb.active
ws.title = "Jira AppSec Matrix"

# Styles
bold = Font(bold=True)
italic = Font(italic=True)
center = Alignment(horizontal="center", vertical="center")

thin = Side(style="thin")
full_border = Border(left=thin, right=thin, top=thin, bottom=thin)
top_bottom_border = Border(top=thin, bottom=thin)

# ===================== TITLE ROW =====================
total_columns = len(all_issue_types) + 2
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)

title_cell = ws.cell(row=1, column=1)
title_cell.value = "Security Jiras"
title_cell.font = Font(bold=True, size=14)
title_cell.alignment = center
title_cell.border = full_border

# ===================== HEADER ROW =====================
ws.append(["Status"] + all_issue_types + ["Total"])
for cell in ws[2]:
    cell.font = bold
    cell.alignment = center
    cell.border = full_border

# ===================== HELPERS =====================
def write_row(label, data, label_font, value_font):
    row = [label]
    total = 0

    for it in all_issue_types:
        val = data.get(it, 0)
        row.append(val)
        total += val

    row.append(total)
    ws.append(row)

    r = ws.max_row
    ws.cell(r, 1).font = label_font
    ws.cell(r, 1).border = full_border

    for c in range(2, ws.max_column + 1):
        cell = ws.cell(r, c)
        cell.alignment = center
        cell.font = value_font
        cell.border = full_border

def merged_blank_row():
    r = ws.max_row + 1
    ws.merge_cells(start_row=r, start_column=1,
                   end_row=r, end_column=ws.max_column)
    for c in range(1, ws.max_column + 1):
        ws.cell(r, c).border = top_bottom_border

# ===================== DATA ROWS =====================
write_row("Reviewed", reviewed, bold, bold)
write_row("Accepted", accepted, italic, italic)
write_row("Rejected", rejected, italic, italic)
merged_blank_row()

write_row("Pending", pending, bold, bold)
write_row("Doable", doable, italic, italic)
write_row("Non Doable", nondoable, italic, italic)
merged_blank_row()

write_row("Overall", overall, bold, bold)

# ===================== SAFE COLUMN WIDTH =====================
for col_idx in range(1, ws.max_column + 1):
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = 16

# ===================== SAVE =====================
output_file = "jira_appsec_issue_type_matrix_FINAL_v2.xlsx"
wb.save(output_file)

print(f"Excel generated successfully: {output_file}")
