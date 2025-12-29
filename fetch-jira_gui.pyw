import requests
from collections import defaultdict
import tkinter as tk
from tkinter import messagebox, scrolledtext
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import threading
import datetime

# ===================== CONFIG =====================
JIRA_URL = "https://jira.demo.almworks.com/jira"
SEARCH_API = f"{JIRA_URL}/rest/api/2/search"

JQLS = {
    "Accepted": (
        "filter=11400 AND labels = AppSecurity "
        "AND labels = appsec_validated "
        "AND labels = appsec_approved"
    ),
    "Rejected": (
        "filter=11400 AND labels = AppSecurity "
        "AND labels = appsec_validated "
        "AND labels = appsec_rejected"
    ),
    "Doable": (
        "filter=11400 AND labels = AppSecurity "
        "AND labels not in (appsec_validated) "
        "AND labels not in (appsec_approved, appsec_rejected) "
        "AND status in (Done, Resolved)"
    ),
    "NonDoable": (
        "filter=11400 AND labels = AppSecurity "
        "AND labels not in (appsec_validated) "
        "AND labels not in (appsec_approved, appsec_rejected) "
        "AND status not in (Done, Resolved)"
    )
}

# ===================== GUI APP =====================
class JiraSecurityGUI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Security Jiras – Excel Generator")
        self.geometry("760x520")
        self.resizable(False, False)

        # --------- AUTH ---------
        auth = tk.LabelFrame(self, text="Jira Authentication", padx=10, pady=10)
        auth.pack(fill="x", padx=10, pady=10)

        tk.Label(auth, text="Username:", width=15, anchor="w").grid(row=0, column=0)
        self.username_entry = tk.Entry(auth, width=40)
        self.username_entry.grid(row=0, column=1)

        tk.Label(auth, text="Password:", width=15, anchor="w").grid(row=1, column=0)
        self.password_entry = tk.Entry(auth, width=40, show="*")
        self.password_entry.grid(row=1, column=1)

        tk.Label(self, text="Security Jiras Report Generator",
                 font=("Segoe UI", 16, "bold")).pack(pady=5)

        self.log = scrolledtext.ScrolledText(self, width=90, height=16)
        self.log.pack(padx=10, pady=10)
        self.log.config(state="disabled")

        tk.Button(self, text="Generate Excel Report",
                  font=("Segoe UI", 12, "bold"),
                  command=self.start_generation).pack(pady=10)

    def log_msg(self, msg):
        self.log.config(state="normal")
        self.log.insert(tk.END, msg + "\n")
        self.log.see(tk.END)
        self.log.config(state="disabled")

    def start_generation(self):
        if not self.username_entry.get() or not self.password_entry.get():
            messagebox.showwarning("Missing Credentials", "Enter Jira username and password")
            return
        threading.Thread(target=self.generate_report, daemon=True).start()

    # ===================== JIRA =====================
    def fetch_issue_type_counts(self, jql, label, auth):
        self.log_msg(f"Fetching {label}...")
        counts = defaultdict(int)
        start_at = 0
        max_results = 100

        while True:
            payload = {
                "jql": jql,
                "startAt": start_at,
                "maxResults": max_results,
                "fields": ["issuetype"]
            }

            r = requests.post(SEARCH_API, json=payload, auth=auth)
            r.raise_for_status()
            data = r.json()

            for issue in data.get("issues", []):
                counts[issue["fields"]["issuetype"]["name"]] += 1

            if start_at + max_results >= data.get("total", 0):
                break
            start_at += max_results

        self.log_msg(f"{label} completed.")
        return dict(counts)

    # ===================== MAIN =====================
    def generate_report(self):
        try:
            auth = (self.username_entry.get(), self.password_entry.get())

            accepted = self.fetch_issue_type_counts(JQLS["Accepted"], "Accepted", auth)
            rejected = self.fetch_issue_type_counts(JQLS["Rejected"], "Rejected", auth)
            doable = self.fetch_issue_type_counts(JQLS["Doable"], "Doable", auth)
            nondoable = self.fetch_issue_type_counts(JQLS["NonDoable"], "Non Doable", auth)

            issue_types = sorted(set(accepted) | set(rejected) | set(doable) | set(nondoable))

            reviewed = {k: accepted.get(k, 0) + rejected.get(k, 0) for k in issue_types}
            pending = {k: doable.get(k, 0) + nondoable.get(k, 0) for k in issue_types}
            overall = {k: reviewed[k] + pending[k] for k in issue_types}

            self.create_excel(issue_types, accepted, rejected, doable,
                              nondoable, reviewed, pending, overall)

            messagebox.showinfo("Success", "Excel report generated successfully")

        except requests.exceptions.HTTPError as e:
            messagebox.showerror("Jira Error", str(e))
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # ===================== EXCEL =====================
    def create_excel(self, issue_types, accepted, rejected,
                     doable, nondoable, reviewed, pending, overall):

        wb = Workbook()
        ws = wb.active
        ws.title = "Jira AppSec Matrix"

        bold = Font(bold=True)
        italic = Font(italic=True)
        center = Alignment(horizontal="center", vertical="center")

        thin = Side(style="thin")
        full = Border(left=thin, right=thin, top=thin, bottom=thin)
        sep = Border(top=thin, bottom=thin)

        total_cols = len(issue_types) + 2

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        t = ws.cell(row=1, column=1)
        t.value = "Security Jiras"
        t.font = Font(bold=True, size=14)
        t.alignment = center
        t.border = full

        ws.append(["Status"] + issue_types + ["Total"])
        for c in ws[2]:
            c.font = bold
            c.alignment = center
            c.border = full

        def row(label, data, lf, vf):
            vals = [label]
            tot = 0
            for it in issue_types:
                v = data.get(it, 0)
                vals.append(v)
                tot += v
            vals.append(tot)
            ws.append(vals)
            r = ws.max_row
            ws.cell(r, 1).font = lf
            ws.cell(r, 1).border = full
            for i in range(2, ws.max_column + 1):
                c = ws.cell(r, i)
                c.font = vf
                c.alignment = center
                c.border = full

        def sep_row():
            r = ws.max_row + 1
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ws.max_column)
            for i in range(1, ws.max_column + 1):
                ws.cell(r, i).border = sep

        row("Reviewed", reviewed, bold, bold)
        row("Accepted", accepted, italic, italic)
        row("Rejected", rejected, italic, italic)
        sep_row()
        row("Pending", pending, bold, bold)
        row("Doable", doable, italic, italic)
        row("Non Doable", nondoable, italic, italic)
        sep_row()
        row("Overall", overall, bold, bold)
